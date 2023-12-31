#!/usr/bin/python3

from openpyxl import Workbook
import openpyxl
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment
import schedlr
import json
import os










def initialize_excel_file(filename='7_19.xlsx'):    
    wb.save(filename=filename)


def merge_row(ws_, row, start_col, end_col):
    ws_.merge_cells(f'{start_col}{row}:{end_col}{row}')

    
def merge_rows(ws_, start_row, end_row, start_col, end_col):
    for row in range(start_row, end_row + 1):
        merge_row(ws_, start_row, start_col, end_col)


def merge_area(ws_, start_row, end_row, start_col, end_col):
    for row in range(start_row, end_row + 1):
        ws_.merge_cells(f'{start_col}{start_row}:{end_col}{end_row}')
    
# top_left_cell.fill = PatternFill("solid", fgColor="DDDDDD")
def cell_fill(cell, color):
    cell.fill = PatternFill("solid", fgColor=color)


def paint_area(ws_, color, start_row, end_row, start_col, end_col):
    columns = list(map(chr, range(ord(start_col), ord(end_col)+1)))
    for row in range(start_row, end_row + 1):
        for col in columns:
            cell_fill(ws_[f"{col}{row}"], color)


def paint_row(ws_, color, start_row, start_col, end_col):
    columns = list(map(chr, range(ord(start_col), ord(end_col)+1)))
    for col in columns:
        cell_fill(ws_[f"{col}{start_row}"], color)
        

def set_cell_value(ws, row, col, value):
    ws[f'{col}{row}'] = value

    
# def get_cell(ws, row, col):
#     return ws.cell
    
        
def generate_color_instructions(ws_, start_row, start_col, end_col):
    color = get_color_scheme()
    colors = ['green', 'yellow', 'red']
    info_text = ['Tillgänglig', 'Krock med icke obligatoriskt moment', 'Krock med obligatoriskt moment']
    info_counter = 0
    row_counter = start_row
    for color_ in colors:
        paint_row(ws_, color[color_], row_counter, start_col, end_col)
        set_cell_value(ws_, row_counter, start_col, info_text[info_counter])
        ws_[f'{start_col}{row_counter}'].alignment = Alignment(horizontal='center')
        merge_row(ws_, row_counter, start_col, chr(ord(end_col)-1))
        row_counter += 1
        info_counter += 1
    return row_counter
    
            
def generate_week_schedule(ws_, start_row, schedule, assistants):
    color = get_color_scheme()
    current_row = start_row + 1
    thin_border = Border(
        left=Side(border_style='thin'),
        right=Side(border_style='thin'),
        top=Side(border_style='thin'),
        bottom=Side(border_style='thin'),
    )
    # Generate grey Vecka X cells
    avail = True
    for week in schedule:
        print(week + ':')
        current_column = 1
        cell_fill(ws_.cell(row=current_row, column=current_column), color['grey'])
        # cell_fill(ws_[f'A{current_row}'], color['grey'])
        set_cell_value(ws_, current_row, 'A', f'Vecka {week}')        
        ws_[f'A{current_row}'].alignment = Alignment(horizontal='center')
        current_row += 1
        # Depth + längden på 


        
        
        # Generate light grey activity schedule
        # Date Time Activity Course ExternalInfo (should be dependent on order)
        for date in schedule[week]:
            for session_time in schedule[week][date]:
                activity_list = [date, session_time] + schedule[week][date][session_time][0:2]
                current_column = 1
                for activity in activity_list:
                    cell_fill(ws_.cell(row=current_row, column=current_column), color['grey'])
                    ws_.cell(row=current_row, column=current_column).value = activity
                    ws_.cell(row=current_row, column=current_column).alignment = Alignment(horizontal='center')
                    ws_.cell(row=current_row, column=current_column).border = thin_border
                    current_column += 1
     
                # Paint Green with assistant names
                counter = 1
                # Hack to create "Tillgänglighet" cell
                if avail == True:
                    avail = False
                    availability_row = current_row - 1
                    availability_col = current_column + counter
                    ws_.merge_cells(start_row=availability_row, start_column=availability_col, end_row=availability_row, end_column=availability_col+1)
                    cell_fill(ws_.cell(row=availability_row, column=availability_col), color['white'])
                    ws_.cell(row=availability_row, column=availability_col).value = 'Tillgänglighet'
                    ws_.cell(row=availability_row, column=availability_col).alignment = Alignment(horizontal='center')
                    ws_.cell(row=availability_row, column=availability_col).border = thin_border
                    ws_.cell(row=availability_row, column=availability_col+1).border = thin_border
                    
                for assistant in assistants:
                    cell_fill(ws_.cell(row=current_row, column=current_column + counter), color['green'])
                    ws_.cell(row=current_row, column=current_column + counter).value = assistant
                    ws_.cell(row=current_row, column=current_column + counter).alignment = Alignment(horizontal='center')
                    ws_.cell(row=current_row, column=current_column + counter).border = thin_border
                    counter += 1
                    
                current_row += 1
        current_row += 1
    return current_column


# Need week_activites_count function in schedlr



def get_color_scheme():
    colors = {
        'green' : 'b6d7a8',
        'yellow' : 'f9cb9c',
        'red' : 'ea9999',
        'grey': 'cccccc',
        'lightgreen' : 'e2efda',
        'lightyellow' : 'fff2cc',
        'lightred' : 'fce4d6',
        'lightblue' : 'd9e1f2',
        'lightgrey' : 'e7e6e6',
        'white' : 'ffffff',
    }
    return colors


def main(filename = "tdp007_019_assistent_schema.xlsx", csv="tdp007_019.csv"):
    # Get correctly formated schedule info.
    formatted_csv = "7_19.csv"
    xl = '7_19.xlsx' 
    schedlr.format_timeedit_categories(csv, formatted_csv)
    schedule = schedlr.initialize_week_schedule(formatted_csv)

    #for week in schedule_dict:
        
    
    # Create Excel file
    assistants = schedlr.get_assistants(formatted_csv, 'TDP007') 

    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Schema tdp007 och tdp019"


    # end row after printing colored instructions decides which row week schedule starts from 
    end_instruction_row = generate_color_instructions(ws1, 2, 'F', 'J') 

    print(json.dumps(schedule, indent=2))
    # Create work schedule info
    generate_week_schedule(ws1, end_instruction_row, schedule, assistants)
    
    
    wb.save(filename=xl)
    #initialize_excel_file()
    
    return "0"



if __name__ == '__main__':
    main()
else:
    print("exceller running as module")


