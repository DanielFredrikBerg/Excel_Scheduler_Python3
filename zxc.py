#!/usr/bin/python3

from openpyxl import Workbook
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.styles import colors
from openpyxl.utils import get_column_letter


redFill = PatternFill(start_color='FFFF0000',
                   end_color='FFFF0000',
                   fill_type='solid')

wb = Workbook()

new_filename = 'tdp007.xlsx'

ws1 = wb.active
ws1.title = "range_names"

for row in range(1, 40):
    ws1[f'A{row}'].fill = redFill
    #ws1.append(range(600))

ws2 = wb.create_sheet(title="pi")
ws2['F5'] = 3.14

ws3 = wb.create_sheet(title="Data")


for row in range(10,20):
    for col in range(27,54):
        _ = ws3.cell(column=col, row=row, value="{0}".format(get_column_letter(col)))

print(ws3['AA10'].value)

wb.save(filename = new_filename)
